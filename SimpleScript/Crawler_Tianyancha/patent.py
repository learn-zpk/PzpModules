import argparse
import random
from time import sleep

import browsercookie
import requests
from lxml import etree
import openpyxl


def param_extract_interactive():
    id = input("请输入公司ID:")
    name = input("请输入公司名称:")
    return id, name


if __name__ == '__main__':
    # parser = argparse.ArgumentParser(description='filename list')
    # parser.add_argument('--company-id', dest='id', type=str,
    #                     help='公司ID', required=True)
    # parser.add_argument('--company-name', dest='name', type=str,
    #                     help='公司名称', required=True)
    # args = parser.parse_args()
    #
    # company_id = args.id
    # company_name = args.name
    company_id, company_name = param_extract_interactive()
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("{}专利".format(company_name))
    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
    worksheet.cell(1, 1, value="专利名称")
    worksheet.cell(1, 2, value="申请公布号")
    sheet_idx = 2

    page_idx = 1
    try:
        while (page_idx < 1000):
            print("page", page_idx, worksheet.max_row - 1)
            patent_uri_template = 'https://www.tianyancha.com/pagination/patent.xhtml?ps=10&pn={}&id={}&_='
            url = patent_uri_template.format(page_idx, company_id)
            cj = browsercookie.chrome()
            response = requests.get(url, cookies=cj)
            if response.status_code != 200:
                raise Exception("爬取失败", page_idx)
            if response.text == '':
                break
            res = etree.HTML(response.text)
            name_list = res.xpath('/html/body/table/tbody/tr/td[3]/span/text()')
            no_list = res.xpath('/html/body/table/tbody/tr/td[5]/span/text()')
            for x in zip(name_list, no_list):
                worksheet.cell(sheet_idx, 1, str(x[0]))
                worksheet.cell(sheet_idx, 2, str(x[1]))
                sheet_idx += 1
            sleep(random.randint(1, 3))
            page_idx += 1
    finally:
        workbook.save("{}-专利.xlsx".format(company_name))
