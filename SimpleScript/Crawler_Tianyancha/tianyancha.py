import argparse
import random

import browsercookie
import openpyxl
import requests
from lxml import etree
from time import sleep


def param_extract_interactive():
    typo_int = int(input("选择类型(专利输入1,软著输入2):"))
    if typo_int == 1:
        typo = '专利'
    elif typo_int == 2:
        typo = '软著'
    else:
        raise Exception("选择类型不支持")
    id = input("请输入公司ID:").strip()
    if not id:
        raise Exception("公司ID不可为空")
    name = input("请输入公司名称:").strip()
    if not name:
        raise Exception("公司名称不可为空")
    print(id, typo, name)
    return typo, id, name


def param_extract_args():
    parser = argparse.ArgumentParser(description='filename list')
    parser.add_argument('--type', dest='typo', type=str,
                        help='类型,支持软著、专利', required=True)
    parser.add_argument('--name', dest='name', type=str,
                        help='公司名称', required=True)
    parser.add_argument('--id', dest='id', type=str,
                        help='公司ID', required=True)
    parser.add_argument('--name', dest='name', type=str,
                        help='公司名称', required=True)
    args = parser.parse_args()
    return args.typo, args.id, args.name


def obtain_response_text(url, page_idx):
    cj = browsercookie.chrome()
    response = requests.get(url, cookies=cj)
    print(response.status_code)
    if response.status_code != 200:
        raise Exception("第{}爬取失败,结束任务".format(page_idx))
    return response.text


def obtain_patents(company_id):
    tmp_result_list = []
    page_idx = 1
    while page_idx < 10000:
        print("page", page_idx, len(tmp_result_list))
        patent_uri_template = 'https://www.tianyancha.com/pagination/patent.xhtml?ps=10&pn={}&id={}&_='
        url = patent_uri_template.format(page_idx, company_id)
        response_text = obtain_response_text(url, page_idx)
        if response_text == '':
            break
        res = etree.HTML(response_text)
        name_list = res.xpath('/html/body/table/tbody/tr/td[3]/span/text()')
        no_list = res.xpath('/html/body/table/tbody/tr/td[5]/span/text()')
        tmp_result_list += zip(name_list, no_list)

        sleep(random.randint(1, 2))
        page_idx += 1
    return tmp_result_list


def obtain_copyrights(company_id):
    tmp_result_list = []
    page_idx = 1
    while page_idx < 10000:
        print("page", page_idx, len(tmp_result_list))
        patent_uri_template = 'https://www.tianyancha.com/pagination/copyright.xhtml?ps=10&pn={}&id={}&_='
        url = patent_uri_template.format(page_idx, company_id)
        response_text = obtain_response_text(url, page_idx)
        if response_text == '':
            break
        res = etree.HTML(response_text)
        detail_name_list = res.xpath('/html/body/table/tbody/tr/td[3]/span/text()')
        simple_name_list = res.xpath('/html/body/table/tbody/tr/td[4]/span/text()')
        version_list = res.xpath('/html/body/table/tbody/tr/td[7]/span/text()')
        tmp_result_list += zip(detail_name_list, simple_name_list, version_list)
        sleep(random.randint(1, 3))
        page_idx += 1
    return tmp_result_list


def write_patent(result_list, company_name, workbook):
    worksheet = workbook.create_sheet("{}专利".format(company_name))
    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
    worksheet.cell(1, 1, value="专利名称")
    worksheet.cell(1, 2, value="申请公布号")
    sheet_idx = 2

    for x in result_list:
        worksheet.cell(sheet_idx, 1, str(x[0]))
        worksheet.cell(sheet_idx, 2, str(x[1]))
        sheet_idx += 1


def write_copyright(result_list, company_name, workbook):
    worksheet = workbook.create_sheet("{}软著".format(company_name))
    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
    worksheet.cell(1, 1, value="软著全称")
    worksheet.cell(1, 2, value="软著简称")
    worksheet.cell(1, 3, value="版本号")
    sheet_idx = 2
    for x in result_list:
        worksheet.cell(sheet_idx, 1, str(x[0]))
        worksheet.cell(sheet_idx, 2, str(x[1]))
        worksheet.cell(sheet_idx, 3, str(x[2]))
        sheet_idx += 1


if __name__ == '__main__':
    # typo,company_id, company_name = param_extract_args()
    typo, company_id, company_name = param_extract_interactive()
    workbook = openpyxl.Workbook()
    if typo == '专利':
        write_patent(obtain_patents(company_id), company_name, workbook)
    elif typo == '软著':
        write_copyright(obtain_copyrights(company_id), company_name, workbook)
    workbook.save("{}-{}.xlsx".format(company_name, typo))
