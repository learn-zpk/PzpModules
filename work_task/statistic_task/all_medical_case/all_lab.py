import collections
import json
import os

import openpyxl

from data.dicts.medical_case_dict import DISEASE_MAP_360S
from utils.pzp_utils import load_data


def process_labs(labs, labs_dict):
    if 'labs' in labs:
        labs = labs['labs']
    for _ in labs:
        try:
            for __ in _['data']:
                item_name = __['name']
                if 'contents' not in __:
                    continue
                for content in __['contents']:
                    sub_item_name = content['normal_item_name']
                    labs_dict[item_name][sub_item_name]['total'] += 1
                    if content['is_abnormal']:
                        labs_dict[item_name][sub_item_name]['abnormal_total'] += 1
                    else:
                        labs_dict[item_name][sub_item_name]['normal_total'] += 1
                    if not labs_dict[item_name][sub_item_name]['normal_sample']:
                        labs_dict[item_name][sub_item_name]['normal_sample'] = content['normal_sample']
                    if not labs_dict[item_name][sub_item_name]['unit']:
                        labs_dict[item_name][sub_item_name]['unit'] = content['unit']
                    if not labs_dict[item_name][sub_item_name]['range_text']:
                        labs_dict[item_name][sub_item_name]['range_text'] = content['range_text']
                    labs_dict[item_name][sub_item_name]['qualitative_result'].add(content['qualitative_result'])
        except Exception as e:
            print(e)
            continue


all_total = 0
labs_total = 0


def process_data(systems, icd10, input_path):
    global all_total, labs_total
    for idx, line in enumerate(load_data(input_path)):
        try:
            medical_record = json.loads(line)
            labs = medical_record['labs']
            all_total += 1
            info_dict[systems][icd10]['total'] += 1
            if not labs:
                continue
            labs_total += 1
            info_dict[systems][icd10]['labs_total'] += 1
            process_labs(labs, info_dict[systems][icd10]['labs'])
        except Exception as e:
            continue


if __name__ == '__main__':
    _input_dir = '/Users/learnzpk/Myspace/pzp_module/work_task/data/output/case_origin'
    _output_dir = '/Users/learnzpk/Myspace/pzp_module/work_task/do_diagnosis/filter_good_medical_case/statistic/统计结果'
    # _input_dir = '/myhome/emrs/case_origin'
    info_dict = collections.defaultdict(lambda: collections.defaultdict(lambda: {}))

    for systems in os.listdir(_input_dir):
        next_input_dir = os.path.join(_input_dir, systems)
        for icd10_x in os.listdir(next_input_dir):
            input_path = os.path.join(next_input_dir, icd10_x)
            icd10 = icd10_x.split('_')[0]
            if not info_dict[systems][icd10]:
                print(systems, icd10)
                info_dict[systems][icd10] = {
                    'total': 0,
                    'labs_total': 0,
                    'labs': collections.defaultdict(lambda: collections.defaultdict(lambda: {
                        'total': 0,
                        'abnormal_total': 0,
                        'normal_total': 0,
                        'normal_sample': '',
                        'unit': '',
                        'range_text': '',
                        'qualitative_result': set()
                    }))
                }
            process_data(systems, icd10, input_path)
    workbook = openpyxl.Workbook()
    worksheet0 = workbook.create_sheet(index=0, title="化验数据统计结果汇总结果")
    worksheet0.cell(1, 1).value = '样本总数'
    worksheet0.cell(1, 2).value = all_total
    worksheet0.cell(1, 3).value = '化验总数'
    worksheet0.cell(1, 4).value = labs_total
    worksheet0.cell(2, 1).value = '系统名'
    worksheet0.cell(2, 2).value = '诊断ICD10'
    worksheet0.cell(2, 3).value = '诊断名'
    worksheet0.cell(2, 4).value = '数目'
    worksheet0.cell(2, 5).value = '化验数目'
    worksheet0.cell(2, 6).value = '化验病历占比'

    i = 3
    for systems, _ in info_dict.items():
        cur_workbook = openpyxl.Workbook()
        index = 0
        for icd10, __ in _.items():
            icd10_name = DISEASE_MAP_360S[icd10]["icd10_name"]
            worksheet0.cell(i, 1).value = systems
            worksheet0.cell(i, 2).value = icd10
            worksheet0.cell(i, 3).value = icd10_name
            worksheet0.cell(i, 4).value = __['total']
            worksheet0.cell(i, 5).value = __['labs_total']
            worksheet0.cell(i, 6).value = 0 if __['total'] == 0 else round(__['labs_total'] / __['total'], 3)
            i += 1

            cur_work_sheet = cur_workbook.create_sheet(
                index=index,
                title=icd10_name.replace('[', '(').replace(']', ')').replace('/', '_'))
            index += 1
            cur_work_sheet.cell(1, 1).value = 'ICD10'
            cur_work_sheet.cell(1, 2).value = '化验名'
            cur_work_sheet.cell(1, 3).value = '样本名'
            cur_work_sheet.cell(1, 4).value = '项目名称'
            cur_work_sheet.cell(1, 5).value = '结果取值'
            cur_work_sheet.cell(1, 6).value = '参考范围'
            cur_work_sheet.cell(1, 7).value = '单位'
            cur_work_sheet.cell(1, 8).value = '异常数目'
            cur_work_sheet.cell(1, 9).value = '正常数目'
            cur_work_sheet.cell(1, 10).value = '样本总数'
            j = 2

            for item_name, xx in __['labs'].items():
                for sub_item_name, xxx in xx.items():
                    cur_work_sheet.cell(j, 1).value = icd10
                    cur_work_sheet.cell(j, 2).value = item_name
                    cur_work_sheet.cell(j, 3).value = xxx.get('normal_sample')
                    cur_work_sheet.cell(j, 4).value = sub_item_name
                    cur_work_sheet.cell(j, 5).value = '/'.join(xxx.get('qualitative_result'))
                    cur_work_sheet.cell(j, 6).value = xxx.get('range_text')
                    cur_work_sheet.cell(j, 7).value = xxx.get('unit')
                    cur_work_sheet.cell(j, 8).value = xxx.get('abnormal_total')
                    cur_work_sheet.cell(j, 9).value = xxx.get('normal_total')
                    cur_work_sheet.cell(j, 10).value = xxx.get('total')
                    j += 1

        cur_workbook.save('{}/{}.xlsx'.format(_output_dir, systems))
    workbook.save('{}/化验数据统计结果汇总.xlsx'.format(_output_dir, ))
